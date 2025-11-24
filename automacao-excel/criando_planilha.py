from openpyxl import Workbook

#isso cria uma nova planilha
wb = Workbook()

ws = wb.active#seleciona a aba ativa

ws.title = "Dados"#da um nome para a aba

'''ws["a1"] = "nome"
ws["b1"] = "email"
ws["c1"] = "telefone"#escreve cabaeçalhos

ws.append(["joão silva", "joao@gmail.com", "11988887777"])
ws.append(["maria silva", "maria@gmail.com", "11988887777"])
ws.append(["josé silva", "jose@gmail.com", "11988887777"])'''#adiciona linhas

ws.append(["id", "nome", "idade", "cidade"])

dados = [
    [1, "ana", 25, "são paulo"],
    [2, "Bruno", 30, "Rio de Janeiro"],
    [3,"Carla", 28, "Belo Horizonte"],
    [4, "Daniel", 35, "curitiba"],
]

for linha in dados:
    ws.append(linha)

wb.save("dados.xlsx")

print("planilha criada com sucesso")

